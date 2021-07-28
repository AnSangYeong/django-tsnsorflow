from bs4 import BeautifulSoup
import requests
import datetime
import urllib
import os
import openpyxl
import re
import pandas as pd

now = datetime.datetime.today()
date = str(now.date()).replace("-", "")

def create_soup(url):
    res = requests.get(url)
    soup = BeautifulSoup(res.text, "lxml")
    return soup

def print_news(rdate, rtime, index, title, press, content, link):
    print("{}. {}. {}. {}. {} ".format(rdate, rtime, index+1, title, press))
    print("{}".format(content))
    print("{}".format(link))
    print("")

def scrape_it_news(date_input):

    print("[IT뉴스]")
    date = urllib.parse.quote(date_input)
    url = "https://news.daum.net/breakingnews/digital?page=1&regDate=" + date
    soup = create_soup(url)
    news_list = soup.find_all("div", attrs={"class":"cont_thumb"})[:15]

    links=[]; titles=[]; contents=[]; agencies=[]; reporting_dates=[]; reporting_times=[];

    for index, news in enumerate(news_list):
        m_idx = 0
        img = news.find("img")
        if img:
            m_idx = 1

        rdate = search_date
        reporting_dates.append(rdate)

        rtime = news.find_all("span", attrs={"class":"info_time"})[m_idx].get_text().strip()
        reporting_times.append(rtime)

        title = news.find_all("a")[m_idx].get_text().strip()
        titles.append(title)

        press = news.find_all("span", attrs={"class":"info_news"})[m_idx].get_text()[0:-8].strip()
        agencies.append(press)

        content = news.find_all("span", attrs={"class":"link_txt"})[m_idx].get_text().strip()
        contents.append(content)

        link = news.find_all("a")[m_idx]["href"]
        links.append(link)

        print_news(rdate, rtime, index, title, press, content, link)

    result = {'link':links, 'title':titles,'content':contents, 'agency':agencies, 'rdate':reporting_dates, 'rtime':reporting_times}
    return result


#    news.loc = [rdate, rtime, index, title, press, content, link]
    print()




if __name__ == "__main__" :
    search_date = input("날짜입력20210101 : ")
    day_news = scrape_it_news(search_date)

    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title=search_date)
#    wb.remove(wb['sheet'])

    col_nums = {}
    for i, k in enumerate(day_news.keys()):
        col_nums[k] = i + 1

    for k in col_nums:
        for row_num in range(1, len(day_news[k])+1):
            if row_num == 1:
                ws.cell(row=row_num, column=col_nums[k]).value = k
            else:
                ws.cell(row=row_num, column=col_nums[k]).value = day_news[k][row_num-1]


    wb.save(os.path.join(os.getcwd(), "dailynews", "daum_news_list_select_day_to_excel.xlsx"))
